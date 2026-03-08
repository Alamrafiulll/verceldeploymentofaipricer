"""Generate a sample product catalog Excel file for bulk import testing."""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)


def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["sku", "name", "category", "unit_cost", "list_price"]
    ws.append(headers)

    products = [
        ("WH-PLATZ-50", "PLATZ DC Pump Water Heater 50L", "water_heater", 920.00, 1499.00),
        ("WH-QUATEK-68", "QUATEK Water Heater 68L", "water_heater", 980.00, 1599.00),
        ("WH-STARKER-100", "STARKER Water Heater 100L", "water_heater", 1250.00, 1999.00),
        ("WH-EDGE-80", "EDGE Water Heater 80L Premium", "water_heater", 1080.00, 1799.00),
        ("WH-STIQ-55", "STIQ Water Heater 55L Economy", "water_heater", 680.00, 1099.00),
        ("HB-ZETA-02", "ZETA Hand Bidet 02 Premium", "hand_bidet", 180.00, 349.00),
        ("HB-ZETA-03", "ZETA Hand Bidet 03 Auto", "hand_bidet", 220.00, 429.00),
        ("PP-TURBO-1", "TURBO Water Pump 1HP", "pump", 450.00, 799.00),
        ("PP-TURBO-1.5", "TURBO Water Pump 1.5HP", "pump", 580.00, 999.00),
        ("PP-FLOW-2", "FLOW Water Pump 2HP Industrial", "pump", 750.00, 1299.00),
        ("AC-SHOWER-SET", "Premium Shower Set Bundle", "accessories", 120.00, 249.00),
        ("AC-FITTING-KIT", "Universal Fitting Kit", "accessories", 35.00, 79.00),
        ("AC-FILTER-3S", "3-Stage Water Filter", "accessories", 180.00, 399.00),
        ("WH-FLUSSO-45", "FLUSSO Instant Water Heater", "water_heater", 420.00, 699.00),
        ("CE-OPC-50KG", "OPC Cement 50kg", "cement", 18.50, 28.00),
    ]

    for row in products:
        ws.append(row)

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    output_path = Path(__file__).parent / "sample_product_catalog.xlsx"
    wb.save(str(output_path))
    print(f"Sample catalog saved to: {output_path}")
    print(f"Total products: {len(products)}")


if __name__ == "__main__":
    generate()
