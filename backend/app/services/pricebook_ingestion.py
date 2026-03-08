import uuid
import csv
import io
from datetime import datetime
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import PolicyDocument, PriceBook, PriceBookItem, Product, User
from app.schemas.policy import PriceBookItemInput, PriceBookUploadRequest
from app.services.audit_logger import log_audit


def _validate_effective_window(
    effective_start: datetime | None,
    effective_end: datetime | None,
) -> None:
    if effective_start and effective_end and effective_end <= effective_start:
        raise ValueError("effective_end must be later than effective_start")


def ingest_pricebook(db: Session, payload: PriceBookUploadRequest, uploaded_by: User) -> PriceBook:
    _validate_effective_window(payload.effective_start, payload.effective_end)
    source_document_uuid: uuid.UUID | None = None
    if payload.source_document_id:
        try:
            source_document_uuid = uuid.UUID(payload.source_document_id)
        except ValueError as exc:
            raise ValueError("Invalid source_document_id") from exc
        exists = db.get(PolicyDocument, source_document_uuid)
        if not exists:
            raise ValueError("Source policy document not found")
    if not payload.items:
        raise ValueError("Pricebook must include at least one item")

    book = PriceBook(
        name=payload.name,
        channel=payload.channel,
        currency=payload.currency,
        effective_start=payload.effective_start,
        effective_end=payload.effective_end,
        source_document_id=source_document_uuid,
        uploaded_by_user_id=uploaded_by.id,
    )
    db.add(book)
    db.flush()

    seen: set[uuid.UUID] = set()
    for item in payload.items:
        try:
            product_id = uuid.UUID(item.product_id)
        except ValueError as exc:
            raise ValueError(f"Invalid product_id: {item.product_id}") from exc

        if product_id in seen:
            raise ValueError("Duplicate product_id in price book items")
        seen.add(product_id)

        product = db.scalar(select(Product).where(Product.id == product_id))
        if not product:
            raise ValueError(f"Product not found for id: {item.product_id}")

        db.add(
            PriceBookItem(
                price_book_id=book.id,
                product_id=product_id,
                list_price=item.list_price,
                notes=item.notes,
            )
        )

    log_audit(
        db=db,
        actor_user_id=str(uploaded_by.id),
        action="pricebook_uploaded",
        entity_type="price_book",
        entity_id=str(book.id),
        new_json={
            "name": payload.name,
            "channel": payload.channel.value,
            "currency": payload.currency,
            "item_count": len(payload.items),
        },
    )
    db.commit()
    db.refresh(book)
    return book


def ingest_pricebook_from_csv(
    db: Session,
    uploaded_by: User,
    file_bytes: bytes,
    name: str,
    channel: str,
    currency: str = "RM",
    effective_start: datetime | None = None,
    effective_end: datetime | None = None,
    source_document_id: str | None = None,
) -> PriceBook:
    items = _parse_pricebook_rows(
        db=db,
        rows=_parse_csv_rows(file_bytes),
    )

    payload = PriceBookUploadRequest(
        name=name,
        channel=channel,
        currency=currency,
        source_document_id=source_document_id,
        effective_start=effective_start,
        effective_end=effective_end,
        items=items,
    )
    return ingest_pricebook(db=db, payload=payload, uploaded_by=uploaded_by)


def _parse_csv_rows(file_bytes: bytes) -> list[dict[str, str]]:
    try:
        decoded = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV file must be UTF-8 encoded") from exc

    reader = csv.DictReader(io.StringIO(decoded))
    required = {"sku", "list_price"}
    headers = {name.strip().lower() for name in (reader.fieldnames or [])}
    if not required.issubset(headers):
        raise ValueError("CSV must include headers: sku,list_price (optional: notes)")
    return [
        {key.strip().lower(): (value or "").strip() for key, value in row.items() if key}
        for row in reader
    ]


def _parse_xlsx_rows(file_bytes: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ValueError("Excel parsing dependency missing. Install openpyxl.") from exc

    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    rows: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        iter_rows = sheet.iter_rows(values_only=True)
        header_map: dict[int, str] | None = None
        for row in iter_rows:
            values = [("" if value is None else str(value).strip()) for value in row]
            if not any(values):
                continue
            if header_map is None:
                normalized = [value.lower() for value in values]
                if "sku" not in normalized or "list_price" not in normalized:
                    break
                header_map = {idx: col for idx, col in enumerate(normalized)}
                continue
            mapped: dict[str, str] = {}
            for idx, cell in enumerate(values):
                column = header_map.get(idx)
                if column:
                    mapped[column] = cell
            if any(mapped.values()):
                rows.append(mapped)

    if not rows:
        raise ValueError("Excel must include a header row with sku,list_price (optional: notes)")
    return rows


def _parse_pricebook_rows(db: Session, rows: list[dict[str, str]]) -> list[PriceBookItemInput]:
    items: list[PriceBookItemInput] = []
    for row in rows:
        sku = row.get("sku", "").strip()
        if not sku:
            continue
        product = db.scalar(select(Product).where(Product.sku == sku))
        if not product:
            raise ValueError(f"SKU not found: {sku}")
        try:
            list_price = float(row.get("list_price", "").strip())
        except ValueError as exc:
            raise ValueError(f"Invalid list_price for SKU {sku}") from exc
        items.append(
            PriceBookItemInput(
                product_id=str(product.id),
                list_price=list_price,
                notes=row.get("notes", "").strip() or None,
            )
        )
    if not items:
        raise ValueError("File contained no valid rows")
    return items


def ingest_pricebook_from_file(
    db: Session,
    uploaded_by: User,
    file_bytes: bytes,
    filename: str,
    name: str,
    channel: str,
    currency: str = "RM",
    effective_start: datetime | None = None,
    effective_end: datetime | None = None,
    source_document_id: str | None = None,
) -> PriceBook:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv_rows(file_bytes)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        rows = _parse_xlsx_rows(file_bytes)
    elif suffix == ".xls":
        raise ValueError("Legacy .xls is not supported. Please convert to .xlsx.")
    else:
        raise ValueError("Unsupported pricebook file. Upload .csv or .xlsx")

    items = _parse_pricebook_rows(db=db, rows=rows)
    payload = PriceBookUploadRequest(
        name=name,
        channel=channel,
        currency=currency,
        source_document_id=source_document_id,
        effective_start=effective_start,
        effective_end=effective_end,
        items=items,
    )
    return ingest_pricebook(db=db, payload=payload, uploaded_by=uploaded_by)
