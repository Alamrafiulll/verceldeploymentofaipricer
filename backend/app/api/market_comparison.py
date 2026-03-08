"""Uploaded-data-first market comparison API."""

from __future__ import annotations

import csv
import io
import uuid

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.deps import get_db, require_roles
from app.db.models import CompetitorProduct, Product, ProductValueProfile, RoleEnum, UploadedFile, User
from app.services.audit_logger import log_audit
from app.services.market_comparison_engine import analyze_product_market_position, update_product_value_profile

router = APIRouter(prefix="/market", tags=["market-comparison"])


def _parse_rows_from_file(file_bytes: bytes, extension: str) -> list[dict]:
    rows: list[dict] = []
    if extension == "csv":
        reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig")))
        return [dict(row) for row in reader]
    if extension == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(value).strip() if value is not None else "" for value in row])))
        return rows
    raise ValueError("Only CSV and XLSX files are accepted")


def _find_product_match(db: Session, category: str, product_name: str, sku: str | None) -> Product | None:
    if sku:
        product = db.scalar(select(Product).where(Product.sku == sku.strip().upper()))
        if product:
            return product
    exact_name = db.scalar(select(Product).where(Product.name == product_name).limit(1))
    if exact_name:
        return exact_name
    category_match = db.scalar(select(Product).where(Product.category == category).limit(1))
    return category_match


@router.post("/competitor-import")
async def import_competitor_data(
    file: UploadFile = File(...),
    source_uploaded_file_id: str | None = Form(default=None),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(RoleEnum.sales, RoleEnum.admin)),
):
    filename = file.filename or "unknown"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are accepted")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File is empty")

    uploaded_file = None
    if source_uploaded_file_id:
        try:
            uploaded_file = db.get(UploadedFile, uuid.UUID(source_uploaded_file_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid uploaded file id") from exc
        if uploaded_file is None:
            raise HTTPException(status_code=404, detail="Uploaded file not found")

    try:
        rows = _parse_rows_from_file(file_bytes=file_bytes, extension=extension)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {exc}") from exc

    required = {"competitor_name", "product_name", "category", "price"}
    if rows:
        missing = required - {str(key).strip().lower() for key in rows[0].keys()}
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {sorted(missing)}")

    imported = 0
    skipped = 0
    errors: list[str] = []
    updated_products: set[str] = set()

    for row_index, row in enumerate(rows, start=2):
        try:
            price = float(str(row.get("price", "0")).replace(",", ""))
            competitor_name = str(row.get("competitor_name", "")).strip()
            product_name = str(row.get("product_name", "")).strip()
            category = str(row.get("category", "")).strip()
            if not competitor_name or not product_name or not category or price <= 0:
                skipped += 1
                continue

            features = {
                "features": str(row.get("features", "")).strip() or None,
                "feature_count": row.get("feature_count") or row.get("features_count"),
                "warranty_months": row.get("warranty_months"),
                "brand_tier": str(row.get("brand_tier", "")).strip() or None,
                "specification_score": row.get("specification_score"),
                "currency": str(row.get("currency", "")).strip() or "RM",
            }
            features = {key: value for key, value in features.items() if value not in (None, "")}

            matched = _find_product_match(
                db=db,
                category=category,
                product_name=product_name,
                sku=str(row.get("matched_sku", "")).strip() or None,
            )
            competitor = CompetitorProduct(
                competitor_name=competitor_name,
                product_name=product_name,
                category=category,
                price=price,
                currency=str(features.get("currency") or "RM"),
                features_json=features,
                source_uploaded_file_id=uploaded_file.id if uploaded_file else None,
                matched_product_id=matched.id if matched else None,
            )
            db.add(competitor)
            imported += 1
            if matched:
                updated_products.add(str(matched.id))
        except Exception as exc:
            errors.append(f"Row {row_index}: {exc}")

    if imported:
        db.flush()
        for product_id in updated_products:
            update_product_value_profile(db=db, product_id=product_id)
        log_audit(
            db=db,
            actor_user_id=str(user.id),
            action="competitor_data_imported",
            entity_type="competitor_product",
            entity_id="batch",
            new_json={
                "imported": imported,
                "skipped": skipped,
                "errors": len(errors),
                "updated_profiles": len(updated_products),
            },
        )
        db.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "updated_profiles": len(updated_products),
        "errors": errors[:10],
        "message": f"Imported {imported} competitor observations and refreshed {len(updated_products)} product value profiles.",
    }


@router.get("/compare/{product_id}")
def compare_product(
    product_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(RoleEnum.sales, RoleEnum.admin, RoleEnum.approver, RoleEnum.executive)),
):
    analysis = analyze_product_market_position(db=db, product_id=product_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Product not found")

    return {
        "product_id": analysis.product_id,
        "product_name": analysis.product_name,
        "category": analysis.category,
        "chin_hin_price": analysis.chin_hin_price,
        "competitor_count": analysis.competitor_count,
        "avg_competitor_price": analysis.avg_competitor_price,
        "min_competitor_price": analysis.min_competitor_price,
        "max_competitor_price": analysis.max_competitor_price,
        "price_gap_percent": analysis.price_gap_percent,
        "positioning_label": analysis.positioning_label,
        "value_positioning_label": analysis.value_positioning_label,
        "value_score": analysis.value_score,
        "recommendation_confidence": analysis.recommendation_confidence,
        "recommended_strategy": analysis.recommended_strategy,
        "market_comparison_summary": analysis.market_comparison_summary,
        "reasoning": analysis.reasoning,
        "matches": [
            {
                "competitor_id": match.competitor_id,
                "competitor_name": match.competitor_name,
                "product_name": match.product_name,
                "competitor_price": match.competitor_price,
                "price_gap_rm": match.price_gap_rm,
                "price_gap_percent": match.price_gap_percent,
                "position": match.position,
                "match_score": match.match_score,
                "quality_proxy_score": match.quality_proxy_score,
                "warranty_months": match.warranty_months,
                "feature_count": match.feature_count,
                "brand_tier": match.brand_tier,
                "specification_score": match.specification_score,
            }
            for match in analysis.matches
        ],
    }


@router.get("/competitors")
def list_competitors(
    category: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(RoleEnum.sales, RoleEnum.admin, RoleEnum.approver, RoleEnum.executive)),
):
    query = select(CompetitorProduct).order_by(CompetitorProduct.category, CompetitorProduct.competitor_name)
    if category:
        query = query.where(CompetitorProduct.category == category)

    competitors = list(db.scalars(query.limit(200)).all())
    return [
        {
            "id": str(competitor.id),
            "competitor_name": competitor.competitor_name,
            "product_name": competitor.product_name,
            "category": competitor.category,
            "price": float(competitor.price),
            "currency": competitor.currency,
            "features": competitor.features_json,
            "matched_product_id": str(competitor.matched_product_id) if competitor.matched_product_id else None,
            "source_uploaded_file_id": str(competitor.source_uploaded_file_id) if competitor.source_uploaded_file_id else None,
        }
        for competitor in competitors
    ]


@router.get("/value-profiles")
def list_value_profiles(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(RoleEnum.sales, RoleEnum.admin, RoleEnum.approver, RoleEnum.executive)),
):
    profiles = list(
        db.scalars(
            select(ProductValueProfile).order_by(ProductValueProfile.updated_at.desc()).limit(200)
        ).all()
    )
    return [
        {
            "id": str(profile.id),
            "product_id": str(profile.product_id),
            "value_score": float(profile.value_score) if profile.value_score is not None else None,
            "positioning_label": profile.positioning_label,
            "price_band": profile.price_band,
            "competitor_count": profile.competitor_count,
            "avg_competitor_price": float(profile.avg_competitor_price) if profile.avg_competitor_price is not None else None,
            "price_gap_percent": float(profile.price_gap_percent) if profile.price_gap_percent is not None else None,
            "recommended_strategy": profile.recommended_strategy,
            "analysis_json": profile.analysis_json,
            "updated_at": profile.updated_at,
        }
        for profile in profiles
    ]
